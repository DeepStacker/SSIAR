import csv
import io
import os
import json
import pandas as pd
import numpy as np
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from fastapi.responses import StreamingResponse
from app.database import get_document

HEADERS_EXTRACTED = [
    "Timestamp",
    "क्लास रोल नंबर",
    "कक्षा",
    "जन्म तिथि",
    "लिंग",
    "मैं दूसरों के साथ अच्छा व्यवहार करने की कोशिश करता हूँ। मुझे उनकी भावनाओं का ध्यान रहता है।]",
    "मैं बेचैन हूँ, मैं ज्यादा देर तक स्थिर नहीं रह सकता]",
    "मुझे बहुत सिरदर्द, पेट दर्द या बीमारी होती है।]",
    "मैं आमतौर पर दूसरों के साथ साझा करता हूँ। उदाहरण के लिए खेल, भोजन]",
    "मुझे बहुत गुस्सा आता है और मैं अक्सर अपना आपा खो देता हूँ]",
    "मैं अपनी उम्र के लोगों के साथ रहने की बजाय अकेले रहना पसंद करूँगा]",
    "मैं आमतौर पर वही करता हूँ जो मुझे कहा जाता है।]",
    "मुझे बहुत चिंता होती है।]",
    "अगर कोई दुखी, परेशान या बीमार है तो मैं उसकी मदद करता हूँ]",
    "मैं लगातार बेचैन या छटपटा रहा हूँ]",
    "मेरा एक या उससे अधिक अच्छा दोस्त है]",
    "मैं बहुत लड़ता हूँ। मैं दूसरों से अपनी मर्जी करवा सकता हूँ]",
    "मैं अक्सर दुखी, उदास या रोता रहता हूँ]",
    "मेरी उम्र के अन्य लोग आम तौर पर मुझे पसंद करते हैं]",
    "मेरा ध्यान आसानी से भटक जाता है, मुझे ध्यान केंद्रित करने में कठिनाई होती है।]",
    "मैं नई परिस्थितियों में घबरा जाता हूँ। मैं आसानी से आत्मविश्वास खो देता हूँ।]",
    "मैं छोटे बच्चों के प्रति दयालु हूँ]",
    "मुझ पर अक्सर झूठ बोलने या धोखा देने का आरोप लगाया जाता है।]",
    "दूसरे बच्चे या युवा लोग मुझे सताते हैं या मुझे धमकाते हैं]",
    "मैं अक्सर दूसरों (माता-पिता, शिक्षक, बच्चे) को कहता हूँ कि मैं  उनकी मदद करने के लिए तैयार हूँ]",
    "मैं कुछ भी करने से पहले सोचता हूँ]",
    "मैं घर, स्कूल या अन्य जगहों से ऐसी चीजें ले आता हूँ जो मेरी नहीं हैं]",
    "मैं अपनी उम्र के लोगों की तुलना में अपने से बड़े लोगों के साथ बेहतर ढंग से घुल-मिल जाता हूँ]",
    "मुझे बहुत डर लगता है, मैं आसानी से डर जाता हूँ]",
    "मैं जो काम कर रहा हूँ उसे पूरा करता हूँ। मेरा ध्यान अच्छा रहता है]",
    "क्या आपकी कोई अन्य टिप्पणी या चिंता है?",
    "गणित",
    "विज्ञान",
    "हिंदी"
]

def format_timestamp(iso_str):
    if not iso_str:
        return ""
    try:
        if "+" in iso_str:
            iso_str = iso_str.split("+")[0]
        dt = datetime.fromisoformat(iso_str)
        # Format as M/D/YYYY H:MM:SS
        return f"{dt.month}/{dt.day}/{dt.year} {dt.hour}:{dt.minute:02d}:{dt.second:02d}"
    except Exception:
        return iso_str

def map_gender(val):
    if not val:
        return ""
    val_clean = str(val).strip().lower()
    if val_clean in ("m", "male"):
        return "Male"
    if val_clean in ("f", "female"):
        return "Female"
    return val

def find_metadata_path():
    path1 = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "shared", "metadata", "sdq_metadata.json"))
    if os.path.exists(path1):
        return path1
    path2 = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "shared", "metadata", "sdq_metadata.json"))
    if os.path.exists(path2):
        return path2
    return path1

def load_metadata():
    path = find_metadata_path()
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except Exception:
            return []

def map_val_to_hindi(val):
    if val is None or val == "":
        return ""
    val_str = str(val).strip()
    digits = []
    if val_str.startswith("[") and val_str.endswith("]"):
        try:
            import ast
            parsed = ast.literal_eval(val_str)
            if isinstance(parsed, (list, tuple)):
                digits = [int(x) for x in parsed]
        except Exception:
            pass
    elif "," in val_str:
        try:
            digits = [int(x.strip()) for x in val_str.split(",") if x.strip().isdigit()]
        except Exception:
            pass
    else:
        try:
            digits = [int(float(val_str))]
        except Exception:
            pass
    if not digits:
        return ""
    labels = {
        1: "सच नहीं",
        2: "कुछ हद तक सच है।",
        3: "सही मे सच है।"
    }
    return ", ".join(labels.get(d, str(d)) for d in digits)

def map_hindi_to_score(val_str, is_reverse):
    if not val_str:
        return ""
    clean_val = val_str.strip().replace("।", "")
    mapping = {
        "सच नहीं": 2 if is_reverse else 0,
        "कुछ हद तक सच है": 1,
        "सही मे सच है": 0 if is_reverse else 2,
        "सही में सच है": 0 if is_reverse else 2
    }
    parts = [p.strip() for p in clean_val.split(",")]
    scores = []
    for p in parts:
        if p in mapping:
            scores.append(str(mapping[p]))
    return ", ".join(scores)

def build_export(
    filtered_docs: list,
    lang: str,
    columns: str | None,
    fmt: str,
):
    meta = load_metadata()
    
    if fmt == "csv":
        rows_data = []
        for d in filtered_docs:
            full_doc = get_document(d["id"])
            if not full_doc:
                continue
                
            row_vals = []
            row_vals.append(format_timestamp(full_doc.get("created_at", "")))
            row_vals.append(full_doc.get("roll_number", ""))
            row_vals.append(full_doc.get("class", ""))
            row_vals.append(full_doc.get("dob", ""))
            row_vals.append(map_gender(full_doc.get("gender", "")))
            
            res = full_doc.get("responses") or {}
            for q in range(1, 26):
                raw_v = res.get(f"q{q}", "")
                row_vals.append(map_val_to_hindi(raw_v))
                
            row_vals.append(full_doc.get("remarks", ""))
            
            acad = full_doc.get("academic_scores") or {}
            row_vals.append(acad.get("math_pct", ""))
            row_vals.append(acad.get("science_pct", ""))
            row_vals.append(acad.get("language_pct", ""))
            
            rows_data.append(row_vals)
            
        return _build_csv(HEADERS_EXTRACTED, rows_data)

    return _build_excel(meta, filtered_docs, lang)

def _build_csv(headers: list, rows: list) -> StreamingResponse:
    output = BytesIO()
    output.write(b'\xef\xbb\xbf')
    text_output = io.TextIOWrapper(output, encoding='utf-8', newline='')
    writer = csv.writer(text_output)
    writer.writerow(headers)
    writer.writerows(rows)
    text_output.detach()
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=ssiar_sdq_digitized.csv"}
    )

def _build_excel(meta: list, filtered_docs: list, lang: str) -> StreamingResponse:
    wb = Workbook()
    
    # Sheet 1: Extracted Data
    ws1 = wb.active
    ws1.title = "Extracted Data"
    
    # Sheet 2: Scored Data
    ws2 = wb.create_sheet(title="Scored Data")
    
    ws1.append(HEADERS_EXTRACTED)
    
    # Build Sheet 2 Headers (matches user sample format + domain scores)
    headers_ws2 = list(HEADERS_EXTRACTED[:5])  # Timestamp to Gender
    headers_ws2.extend(HEADERS_EXTRACTED[5:30])  # Questions
    headers_ws2.extend(["score_prosocial", "score_emotional", "score_conduct", "score_hyperactivity", "score_peer", "score_total_difficulties"])
    headers_ws2.extend(HEADERS_EXTRACTED[30:])  # Remarks + Academics
    ws2.append(headers_ws2)
    
    orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
    green_fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    
    for d in filtered_docs:
        full_doc = get_document(d["id"])
        if not full_doc:
            continue
            
        row_status = full_doc.get("status", "")
        conf_scores = full_doc.get("confidence_scores")
        
        review_fields = []
        if conf_scores:
            try:
                conf_data = json.loads(conf_scores) if isinstance(conf_scores, str) else conf_scores
                review_fields = conf_data.get("review_fields", [])
            except Exception:
                pass
                
        # Build Row Maps
        res = full_doc.get("responses") or {}
        acad = full_doc.get("academic_scores") or {}
        
        ws1_row_vals = []
        ws1_row_vals.append(format_timestamp(full_doc.get("created_at", "")))
        ws1_row_vals.append(full_doc.get("roll_number", ""))
        ws1_row_vals.append(full_doc.get("class", ""))
        ws1_row_vals.append(full_doc.get("dob", ""))
        ws1_row_vals.append(map_gender(full_doc.get("gender", "")))
        
        for q in range(1, 26):
            raw_v = res.get(f"q{q}", "")
            ws1_row_vals.append(map_val_to_hindi(raw_v))
            
        ws1_row_vals.append(full_doc.get("remarks", ""))
        ws1_row_vals.append(acad.get("math_pct", ""))
        ws1_row_vals.append(acad.get("science_pct", ""))
        ws1_row_vals.append(acad.get("language_pct", ""))
        
        ws1.append(ws1_row_vals)
        excel_row_num = ws1.max_row
        
        # Apply formatting to Sheet 1 cells (highlighting pending/verified fields)
        for col_idx in range(1, len(ws1_row_vals) + 1):
            cell = ws1.cell(row=excel_row_num, column=col_idx)
            
            # Map index to DB fields
            if col_idx == 1:
                field_name = "created_at"
            elif col_idx == 2:
                field_name = "roll_number"
            elif col_idx == 3:
                field_name = "class"
            elif col_idx == 4:
                field_name = "dob"
            elif col_idx == 5:
                field_name = "gender"
            elif 6 <= col_idx <= 30:
                field_name = f"q{col_idx - 5}"
            elif col_idx == 31:
                field_name = "remarks"
            elif col_idx == 32:
                field_name = "math_pct"
            elif col_idx == 33:
                field_name = "science_pct"
            elif col_idx == 34:
                field_name = "language_pct"
            else:
                field_name = ""
                
            if row_status == "verified":
                cell.fill = green_fill
            elif row_status == "needs_review" and field_name in review_fields:
                cell.fill = orange_fill
                
        # Populate Sheet 2 (Scored Values)
        scored_q_vals = {}
        for q in range(1, 26):
            meta_item = next((item for item in meta if item["question_id"] == f"q{q}"), None)
            is_reverse = meta_item.get("reverse_scored", False) if meta_item else False
            scored_q_vals[f"q{q}"] = map_hindi_to_score(ws1_row_vals[q + 4], is_reverse)
            
        # Calculate Domain scores
        domain_scores = {}
        if meta:
            domains = set(item["domain"] for item in meta)
            for dom in domains:
                dom_items = [item["question_id"] for item in meta if item["domain"] == dom]
                row_scores = []
                for q_id in dom_items:
                    q_score_str = scored_q_vals.get(q_id, "")
                    if q_score_str != "":
                        try:
                            val_int = int(q_score_str.split(",")[0])
                            row_scores.append(val_int)
                        except Exception:
                            pass
                if len(row_scores) >= 3:
                    raw_sum = sum(row_scores)
                    scaled_sum = round((raw_sum / len(row_scores)) * 5)
                    domain_scores[f"score_{dom.lower()}"] = scaled_sum
                    
            diff_domains = ["emotional", "conduct", "hyperactivity", "peer"]
            if all(f"score_{x}" in domain_scores for x in diff_domains):
                domain_scores["score_total_difficulties"] = sum(domain_scores[f"score_{x}"] for x in diff_domains)
                
        # Write Sheet 2 Row
        ws2_row_vals = list(ws1_row_vals[:5])  # Timestamp to Gender
        for q in range(1, 26):
            ws2_row_vals.append(scored_q_vals.get(f"q{q}", ""))
        for dom in ["prosocial", "emotional", "conduct", "hyperactivity", "peer", "total_difficulties"]:
            ws2_row_vals.append(domain_scores.get(f"score_{dom}", ""))
        ws2_row_vals.extend(ws1_row_vals[30:])  # Remarks + Academics
        
        ws2.append(ws2_row_vals)
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ssiar_sdq_digitized.xlsx"}
    )
